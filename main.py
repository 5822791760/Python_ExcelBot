import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Font
import re


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def get_start_rows(sheet_object):
    col_num = find_match_key(sheet_object[1], '^inv')
    for key, row in enumerate(sheet_object, -1):
        if row[col_num].value is None:
            return key

    return 0


def find_match_val(df, reg):
    for col in df.columns:
        if re.search(reg, col, flags=re.IGNORECASE):
            return col


def find_match_key(sheet_val, reg):
    for k, v in enumerate(sheet_val):
        if re.search(reg, v.value, flags=re.IGNORECASE):
            return k


def rename_column(df, reg_list, name_list):
    for i in range(len(name_list)):
        df = df.rename(columns={find_match_val(df, reg_list[i]): name_list[i]})
    return df


def get_input():
    global loc_write
    global loc_read

    readfile1 = False
    readfile2 = False

    while not readfile1:
        loc_write = input("\033[94mPlease submit UTL file:\033[0m ").replace("\\", "").replace("\'", "").rstrip()
        try:
            pd.read_excel(loc_write)
            readfile1 = True
        except:
            print("\033[1;91mCan't read file, please insert correct file\033[0m")

    while not readfile2:
        loc_read = input("\033[35mPlease submit Korea Comparison:\033[0m ").replace("\\", "").replace("\'", "").rstrip()
        try:
            pd.read_excel(loc_read)
            readfile2 = True
        except:
            print("\033[1;91mCan't read file, please insert correct file\033[0m")


# =========================================================================

get_input()

utl_all_sheet = pd.read_excel(loc_write, sheet_name=None)
compare_file = pd.read_excel(loc_read)
write_file = load_workbook(loc_write)

sheet_count = len(write_file.sheetnames)

compare_file["Line"] = (compare_file["Line"]/10).astype("int64")

re_list2 = ['^po', '^hds.*e$', '^line', '^pn$', '^p.*2$', '^fg1$', '^fg2$', 'wip']
change_name2 = ["PO", "HDS code", "Line", "#STOCK", "Rev", "FG1", "FG2", "WIP"]

compare_file = rename_column(compare_file, re_list2, change_name2)

utl_file = pd.DataFrame()

re_list = ['po', 'line', 'code', 'stock', 'rev', '^fg', 'wip']
change_name = ["PO", "Line", "HDS code", "#STOCK", "Rev.", "FG", "WIP"]

re_write = ['comment', '^fg', '^wip']

for df in utl_all_sheet.values():
    df = rename_column(df, re_list, change_name)
    utl_file = pd.concat([utl_file, df[change_name]])

utl_file = utl_file.reset_index(drop=True)

compare_file["FG"] = np.where(compare_file["FG1"] != 0, compare_file["FG1"], compare_file["FG2"])
read_file = compare_file[["PO", "HDS code", "Line", "#STOCK", "Rev", "FG", "WIP"]]
read_file = read_file[~read_file["PO"].str.contains("[A-Za-z-]", case=True, regex=True)]
read_file["PO"] = read_file["PO"].astype("int64")

utl_file = utl_file.reset_index()

read_file["Rev"] = read_file["Rev"].apply(lambda x: 0 if (x == "O" or x == "o") else x)

read_file['#STOCK'] = read_file['#STOCK'].astype("|S")
utl_file['#STOCK'] = utl_file['#STOCK'].astype("|S")

real_table = pd.merge(utl_file, read_file, how="left", on=["PO", "#STOCK", "HDS code"], indicator=True)
real_table.drop(real_table[(real_table.duplicated(subset="index")) & (real_table["Line_x"] != real_table["Line_y"])].index, inplace=True)
real_table.drop(real_table[(real_table.duplicated(subset="index", keep=False)) & (real_table["Line_x"] != real_table["Line_y"])].index, inplace=True)
real_table.drop_duplicates(subset="index", inplace=True)
real_table["Comment"] = np.where(real_table["FG_y"].isna(), "CLOSED", np.where(real_table["Line_x"] != real_table["Line_y"],
                                 "Check Line", np.where(real_table["Rev"] != real_table["Rev."], "No Rev", "")))
real_table = real_table.drop(columns=["index"])
real_table = real_table.reset_index(drop=True)
real_table.rename(columns={"Line_x": "Line", "FG_y": "FG", "WIP_y": "WIP"}, inplace=True)
real_table = real_table[["PO", "Line", "HDS code", "#STOCK", "Rev.", "Comment", "FG", "WIP"]]
real_table.loc[real_table["Comment"] == "CLOSED", ["FG", "WIP"]] = ""

min_row = 0
max_row = 0
sheet_count = len(write_file.sheetnames)

for i in range(sheet_count):
    print(f"WRITING {write_file.sheetnames[i]}.....", end="")
    header_idn = []
    active_sheet = write_file.worksheets[i]
    row_count = get_maximum_rows(sheet_object=active_sheet) - 1
    start_row = get_start_rows(active_sheet)
    max_row += row_count
    for re_val in re_write:
        header_idn.append(find_match_key(active_sheet[1], re_val)+1)

    rows = list(dataframe_to_rows(real_table[min_row:max_row], index=False, header=False))
    for r_ind, row in enumerate(rows[start_row:], start=start_row+2):
        active_sheet.cell(row=r_ind, column=header_idn[0], value=row[5]).font = Font(
            color="00FF0000") if row[5] == "CLOSED" else Font(color="000066CC") if row[5] == "Check Line" else Font(color="00969696")
        active_sheet.cell(row=r_ind, column=header_idn[1], value=row[6]).font = Font(color="00969696") if (row[6] == 0 or row[6] == None) else Font(color="00003300")
        active_sheet.cell(row=r_ind, column=header_idn[2], value=row[7]).font = Font(color="00969696") if (row[7] == 0 or row[7] == None) else Font(color="00003300")

    min_row += row_count
    print("COMPLETE")

write_file.save(loc_write)
